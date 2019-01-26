import re
import psycopg2 as psycopg2
from flask import Flask, request, redirect, url_for
from flask import render_template
from openpyxl import load_workbook

app = Flask(__name__)
search_path = 'SET search_path TO public'
count = 0
number = 0
version = 'a82'


class Factory:

    def __init__(self, name, days, item, ingredient1, ingredient_1_required, ingredient2, ingredient_2_required,
                 ingredient3, ingredient_3_required, produce_per_month):
        self.name = name
        self.days = days
        self.item = item
        self.ingredient1 = ingredient1
        self.ingredient_1_required = ingredient_1_required
        self.ingredient_1_factory_required = ""
        self.ingredient2 = ingredient2
        self.ingredient_2_required = ingredient_2_required
        self.ingredient_2_factory_required = ""
        self.ingredient3 = ingredient3
        self.ingredient_3_required = ingredient_3_required
        self.ingredient_3_factory_required = ""
        self.produce_per_month = produce_per_month

    def get_name(self):
        return self.name

    def get_days(self):
        return self.days

    def get_item(self):
        return self.item

    def get_ingredient(self):
        return self.ingredient1

    def get_ingredient_1_required(self):
        return self.ingredient_1_required

    def get_ingredient2(self):
        return self.ingredient2

    def get_ingredient_2_required(self):
        return self.ingredient_2_required

    def get_ingredient3(self):
        return self.ingredient3

    def get_ingredient_3_required(self):
        return self.ingredient_3_required

    def get_produce_per_month(self):
        return self.produce_per_month

    def get_item(self):
        return self.item

    def get_ingredient1(self):
        return self.ingredient1

    def get_ingredient2(self):
        return self.ingredient2

    def get_ingredient3(self):
        return self.ingredient3

    def to_string(self, multiplier):
        global count
        output = self.name + str(self.days) + self.item
        if type(self.ingredient1) == str:
            output = output + self.ingredient1
        else:
            output = output + self.ingredient1.item

        output = output + str(self.ingredient_1_required)

        if type(self.ingredient2) == str:
            output = output + self.ingredient2
        else:
            output = output + self.ingredient2.item

        output = output + str(self.ingredient_2_required)

        if type(self.ingredient3) == str:
            output = output + self.ingredient3
        else:
            output = output + self.ingredient3.item
        output = output + str(self.ingredient_3_required)

        # # print(multiplier * self.produce_per_month)
        if multiplier:
            output = output + str(float(self.produce_per_month) * float(multiplier))
        if type(self.ingredient1) == Factory:
            count = count + 1
            output = output + "\n" + "\t" * count + str(self.ingredient_1_factory_required) + \
                     self.ingredient1.to_string(self.ingredient_1_factory_required)
            count = count - 1
        if type(self.ingredient2) == Factory:
            count = count + 1
            output = output + "\n" + "\t" * count + str(self.ingredient_2_factory_required) + \
                     self.ingredient2.to_string(self.ingredient_2_factory_required)
            count = count - 1
        if type(self.ingredient3) == Factory:
            count = count + 1
            output = output + "\n" + "\t" * count + str(self.ingredient_3_factory_required) + \
                     self.ingredient3.to_string(self.ingredient_3_factory_required)
            count = count - 1

        return output

    # TODO: So, factory toList ingredients are .to_string or 0, if 0 in html, don't add row.


all_factories = []

# TODO: FILES OVER WRITE EACH OTHER UPON BUILDING FACTORY OBJECTS!
def create_factories(item):
    global all_factories
    new_factory = Factory(item[0], item[1], item[2], item[3], item[4], item[5], item[6], item[7], item[8], item[9])
    all_factories.append(new_factory)


def build_table(version):
    wb = load_workbook(filename='resources_' + version + '.xlsx')
    ws = wb.worksheets[0]
    # for sheet in wb:
        # print(sheet)
    j = 1

    for row in ws.iter_rows(min_row=1, max_col=10, max_row=139):
        i = 0
        # Factory, Days, item, Ingredient1, NumIngredient1, Ingredient2, NumIngredient2, Ingredient3, NumIngredient3, MonthProduce
        factory = [None, 0, None, None, 0, None, 0, None, 0, 0.0]
        if row is not None:
            for cell in row:
                if cell.value is None:
                    item = 0
                elif str(cell.value).lower() == 'n/a':
                    item = 0
                else:
                    item = cell.value
                item = re.sub('[^A-Za-z\d./ ]', '', str(item))
                factory[i] = item
                i = i + 1
            if (factory[0] != '') & (factory[2] != 0):
                connstr = "host='localhost' dbname='postgres' user='postgres' password='password'"
                conn = psycopg2.connect(connstr)
                cur = conn.cursor()
                cur.execute(search_path)
                cur.execute(
                    "INSERT INTO factories_" + version + " VALUES ('%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s')" %
                    (factory[0], factory[1], factory[2], factory[3], factory[4], factory[5], factory[6], factory[7],
                     factory[8], factory[9]))
                conn.commit()

            # print(str(j) + "/139")
            j = j + 1


def create_ingredient_nodes():
    for factory in all_factories:
        if factory.ingredient1 != '0':
            for factory_2 in all_factories:
                if factory.ingredient1 == factory_2.item:
                    factory.ingredient1 = factory_2
        if factory.ingredient2 != '0':
            for factory_2 in all_factories:
                if factory.ingredient2 == factory_2.item:
                    factory.ingredient2 = factory_2
        if factory.ingredient3 != '0':
            for factory_2 in all_factories:
                if factory.ingredient3 == factory_2.item:
                    factory.ingredient3 = factory_2


def get_your_data_from_db(version):
    items = []
    connstr = "host='localhost' dbname='postgres' user='postgres' password='password'"
    conn = psycopg2.connect(connstr)
    cur = conn.cursor()
    cur.execute(search_path)
    cur.execute("SELECT * FROM factories_" + version + "")
    test = []
    for item in cur.fetchall():
        test.append(item)
    return test


def get_your_list_data_from_db(version):
    items = []
    connstr = "host='localhost' dbname='postgres' user='postgres' password='password'"
    conn = psycopg2.connect(connstr)
    cur = conn.cursor()
    cur.execute(search_path)
    cur.execute("SELECT item FROM factories_" + version + " ORDER BY item")
    test = []
    for item in cur.fetchall():
        test.append(item[0])
    return test


def find_ratios(item, multiplier):
    if type(item.ingredient1) == Factory:
        child_item1 = item.ingredient1

        required_per_month = ((item.ingredient_1_required / item.days) * 30) * multiplier
        # # print(str(required_per_month) + " = (" + str(item.ingredient_1_required) + " / " + str(item.days) + ") * 30")
        ratio = required_per_month / child_item1.produce_per_month
        # # print(str(ratio) + " = " + " / " + str(required_per_month) + str(child_item1.produce_per_month))
        # Add 1 so is rounded down to upper number when made int, 2.3 would require 3 so 3.3 to int = 3
        if ratio % 1 != 0:
            ratio = ratio + 1
        ratio = int(ratio)
        item.ingredient_1_factory_required = ratio
        find_ratios(child_item1, ratio)

    if type(item.ingredient2) == Factory:
        child_item2 = item.ingredient2

        required_per_month = ((item.ingredient_2_required / item.days) * 30) * multiplier
        # # print(str(required_per_month) + " = (" + str(item.ingredient_2_required) + " / " + str(item.days) + ") * 30")
        ratio = required_per_month / child_item2.produce_per_month
        # # print(str(ratio) + " = " + " / " + str(required_per_month) + str(child_item2.produce_per_month))
        # Add 2 so is rounded down to upper number when made int, 2.3 would require 3 so 3.3 to int = 3
        if ratio % 1 != 0:
            ratio = ratio + 1
        ratio = int(ratio)
        item.ingredient_2_factory_required = ratio

        find_ratios(child_item2, ratio)

    if type(item.ingredient3) == Factory:
        child_item3 = item.ingredient3

        required_per_month = ((item.ingredient_3_required / item.days) * 30) * multiplier
        # # print(str(required_per_month) + " = (" + str(item.ingredient_3_required) + " / " + str(item.days) + ") * 30")
        ratio = required_per_month / child_item3.produce_per_month
        # # print(str(ratio) + " = " + " / " + str(required_per_month) + str(child_item3.produce_per_month))
        # Add 3 so is rounded down to upper number when made int, 3.3 would require 3 so 3.3 to int = 3
        if ratio % 1 != 0:
            ratio = ratio + 1
        ratio = int(ratio)
        item.ingredient_3_factory_required = ratio
        find_ratios(child_item3, ratio)


@app.context_processor
def utility_functions():
    def print_in_console(message):
        print(str(message))

    return dict(mdebug=print_in_console)


def create_data(version):
    # build_table()
    print("VERSION IN DATA: " + str(version))
    items = get_your_data_from_db(version)
    for item in items:
        create_factories(item)
    # for factories in all_factories:
    # # print(factories.to_string())
    create_ingredient_nodes()


def find_factory(item):
    for factory in all_factories:
        if item == factory.item:
            return factory


@app.route('/')
def hello_world():
    return redirect(url_for('profile', item='Wood'))


@app.route('/version', methods=['POST'])
def route():
    global version

    print("VERSION PREVIOUS: " + str(version))
    print("VERSION: " + str(request.form['submit_button']))

    if str(request.form['submit_button']) == 'A8.2':
        version = 'a82'
    elif str(request.form['submit_button']) == 'A9.0':
        version = 'a9'
    print("VERSION AFTER: " + str(version))
    return redirect('http://127.0.0.1:5000/Chicken%20Meat')


@app.route('/<string:item>', methods=['GET', 'POST'])
def profile(item):
    global number, version
    logf = open("download.log", "w")
    try:
        print("IN BUILDER: " + str(version))
        # build_table('a82')
        # build_table('a9')
        # print("Version: " + str(version))
        if request.method == 'POST':
            number = float(request.form['quantity'])
            # print(number)
        items = get_your_list_data_from_db(version)
        create_data(version)
        print(version)
        if number <= 0:
            number = 1
        multiplier = number
        number = int(number)
        factory = find_factory(item)
        find_ratios(factory, multiplier)
        list = [factory]
        # for i in list:
            # print(i)

    except Exception as e:
        logf.write(str(e))
    logf.close()
    return render_template('index.html', items=items, root_factory=factory, multiplier=multiplier,
                           multiplier_int=number)


if __name__ == '__main__':
    app.run()
